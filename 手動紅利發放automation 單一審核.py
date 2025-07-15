import yaml,os
import requests,logging,json
from datetime import datetime
from openpyxl import Workbook
from itertools import cycle

logging.basicConfig(
    level=logging.INFO,
    format='%(asctime)s - %(levelname)s - %(message)s',
    datefmt='%Y-%m-%d %H:%M:%S'
)
class B_end:
    def __init__(self,credential:dict):
        self.session=requests.Session()
        self.username=''
        self.password=''
        self.token=self.get_token(credential['operatorName'],credential['password'])
        self.credential=credential
        self.token_data=self.token
        self.record_data_list=[]
        self.claimid_list=[]
        self.success_count=0
        self.claimid=''
    def get_token(self,operatorName,password):
        login_url="http://sit-admin2.tcg.com/tac/api/login/password"
        payload={
                "operatorName": operatorName,
                "password": password
            }
        headers = {
            "Accept": "application/json, text/plain, */*",
            "Accept-Language": "en-US,en;q=0.9",
            "Authorization": "",
            "Connection": "keep-alive",
            "Content-Type": "application/json",
            "Merchant": "gi8viet",
            "MerchantCode": "gi8viet",
            "Origin": "http://sit-admin2.tcg.com",
            "Referer": "http://sit-admin2.tcg.com/",
            "User-Agent": "Mozilla/5.0 (Macintosh; Intel Mac OS X 10_15_7) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/133.0.0.0 Safari/537.36",
            "environment": "",
            "language": "zh_CN",
            "noErrorNotice": "true",
            "platform": "",
            "Tac-Trace-Id":"&kZEwhHNN!Pe(Qj_"
        }
        
        cookies = {
            "language": "zh_CN",
            "JSESSIONID":"wK3EQfljeUHXxYAN8uKQcvkpKBg1WM4PaVshMx7TpsBoHDtAk4c_!-1653539373"
        }
        requests_data=requests.post(login_url,json=payload,headers=headers,cookies=cookies,verify=False)
        token_data=requests_data.json()
        token=token_data.get("token")
        logging.info(f"登入API回傳: {token}")
        return token

    def create_bonus(self,player:str,bonusAmount:int,bonusPointAmount:int,ticketId:int,ticketQuantity:int,prmotion_id:int):
        
        API_URL = "http://sit-admin2.tcg.com/tac/api/relay/post/mcs-manual-promotion-addManualPromotionClaim" 
        payload = {
        "merchantCode": "gi8viet",
        "customerName": player,
        "bonusAmount": bonusAmount,
        "bonusPointAmount": bonusPointAmount,
        "promotionId": prmotion_id,
        "toReqAmount": 0,
        "ticketId": ticketId,
        "ticketQuantity": ticketQuantity
    }

        headers = {
        "Accept": "application/json, text/plain, */*",
        "Accept-Language": "en-US,en;q=0.9",
        "Authorization": self.token_data,
        "Content-Type": "application/json",
        "Connection": "keep-alive",
        "Language": "zh_CN",
        "Merchant": "gi8viet",
        "Origin": "http://sit-admin2.tcg.com",
        "Referer": "http://sit-admin2.tcg.com/24785",
        "User-Agent": "Mozilla/5.0 (Macintosh; Intel Mac OS X 10_15_7) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/133.0.0.0 Safari/537.36",
        "environment": "TCG3",
        "merchantCode": "gi8viet",
        "platform": "TCG"
        }
        cookies = {
            "language": "zh_CN"
        }
        try:
            response = requests.post(API_URL, json=payload, headers=headers, cookies=cookies, verify=False)
            response.raise_for_status()
            
            
            response_data = response.json()
            logging.info(f"狀態碼: {response.status_code}")
            logging.info(f"響應內容: {response_data}")
            
            
            if response_data.get("success") == True:
                logging.info(f"手動紅利發放成功 ")
                return True
            else:
                error_msg = response_data.get("message", "未知錯誤")
                logging.error(f"手動紅利發放失敗: {error_msg}")
                return False
                
        except requests.RequestException as e:
            logging.error(f"HTTP錯誤 {e}")
            return False
        except ValueError as e:
            logging.error(f"JSON解析錯誤: {e}")
            return False
        except Exception as e:
            logging.error(f"其他錯誤: {e}")
            return False
    def Search_Customer_bonus(self,player:str):
      
        API_URL = "http://sit-admin2.tcg.com/tac/api/relay/get/mcs-manualPromotion-search" 
        start_time = datetime.now().strftime("%Y-%m-%d 00:00:00")
        end_time = datetime.now().strftime("%Y-%m-%d 23:59:59")
        payload = {
        "merchantCode": "gi8viet",
        "status": "P",
        "customerName":player,
        "searchDateMode": "issuedDateSearch",
        "startTime": start_time,
        "endTime": end_time,
        "pageSize": 10,
        "pageNo": 1
    }

        headers = {
        "Accept": "application/json, text/plain, */*",
        "Accept-Language": "en-US,en;q=0.9",
        "Authorization": self.token_data,
        "Content-Type": "application/json",
        "Connection": "keep-alive",
        "Language": "zh_CN",
        "Merchant": "gi8viet",
        "Origin": "http://sit-admin2.tcg.com",
        "Referer": "http://sit-admin2.tcg.com/24785",
        "User-Agent": "Mozilla/5.0 (Macintosh; Intel Mac OS X 10_15_7) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/133.0.0.0 Safari/537.36",
        "environment": "TCG3",
        "merchantCode": "gi8viet",
        "platform": "TCG"
        }
        cookies = {
            "language": "zh_CN"
        }
        try:
            response = requests.get(API_URL, params=payload, headers=headers, cookies=cookies, verify=False)
            response.raise_for_status()
            
            response_data = response.json()
            logging.info(f"狀態碼: {response.status_code}")
            
            
            if response_data.get("success") == True:
                customer_list=response_data.get("value",[])
                
                if not customer_list:
                    logging.error("回應中找不到 customerlist")

                customer_info=customer_list[0]
                CustomerID=customer_info.get("customerId")
                claimid=customer_info.get("id")
        
                if CustomerID and claimid:
                    return CustomerID, claimid
                
            else:
                    logging.error("回應中找不到 customerId 或 claimid")
                    return None, None 
            
                
        except requests.RequestException as e:
            logging.error(f"HTTP錯誤 {e}")
            return None, None
        except ValueError as e:
            logging.error(f"JSON解析錯誤: {e}")
            return None, None
        except Exception as e:
            logging.error(f"其他錯誤: {e}")
            return None, None
    
    def Confirm_Customer_bonus(self,Customerid:int):
    
        API_URL = f"http://sit-admin2.tcg.com/tac/api/relay/post/mcs-manual-promotion-approveClaimStatus?claimStatus=I&customerId={Customerid}&claimId={self.claimid}" 
        payload = {
        "internalRemark": ""
        }

        headers = {
        "Accept": "application/json, text/plain, */*",
        "Accept-Language": "en-US,en;q=0.9",
        "Authorization": self.token_data,
        "Connection": "keep-alive",
        "Language": "zh_CN",
        "Merchant": "gi8viet",
        "Origin": "http://sit-admin2.tcg.com",
        "Referer": "http://sit-admin2.tcg.com/24785",
        "User-Agent": "Mozilla/5.0 (Macintosh; Intel Mac OS X 10_15_7) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/133.0.0.0 Safari/537.36",
        "environment": "TCG3",
        "merchantCode": "gi8viet",
        "platform": "TCG"
        }
        cookies = {
            "language": "zh_CN",
            "JSESSIONID":"wK3EQfljeUHXxYAN8uKQcvkpKBg1WM4PaVshMx7TpsBoHDtAk4c_!-1653539373"
        }
        try:
            response = requests.post(API_URL, json=payload, headers=headers, cookies=cookies, verify=False)
            response_data = response.json()
            
            if response_data.get("success") == True:
                logging.info(f"審核活動紅利成功 ")
                return True
            else:
                error_msg = response_data.get("value" )
                logging.error(f"未審核成功 value: {error_msg}")
                return False
                
        except requests.RequestException as e:
            logging.error(f"HTTP錯誤 {e}")
            return False
        except ValueError as e:
            logging.error(f"JSON解析錯誤: {e}")
            return False
        except Exception as e:
            logging.error(f"其他錯誤: {e}")
            return False
    def Bonus_record_page(self):
        
      
        API_URL2=f"http://sit-admin2.tcg.com/tac/api/relay/get/mcs-v2-promotionClaim-search?pageSize=20&pageNo=1"  
        start_time = datetime.now().strftime("%Y-%m-%d 00:00:00")
        end_time = datetime.now().strftime("%Y-%m-%d 23:59:59")
        headers={
            "Accept": "application/json, text/plain, */*",
            "Accept-Language": "en-US,en;q=0.9",
            "Authorization": self.token_data,
            "Content-Type": "application/json",
            "Connection": "keep-alive",
            "Language": "zh_CN",
            "Merchant": "gi8viet",
            "MerchantCode": "gi8viet",
            "Tac-Trace-Id":"2eAM8QMqpfEd3QxE",
            "Referer": "http://sit-admin2.tcg.com/311792",
            "User-Agent": "Mozilla/5.0 (Macintosh; Intel Mac OS X 10_15_7) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/133.0.0.0 Safari/537.36",
            "environment": "TCG3",
            "merchantCode": "gi8viet",
            "notPending": "true",
            "platform": "TCG"
        }
        payload={
            "fromDate":start_time,
            "toDate":end_time,
            "isFuzzySearch":True,
            "searchDateMode":"requestedTimeSearch",
            "merchantCode":"gi8viet",

        }
        cookies = {
            "language": "zh_CN",
            "JSESSIONID":"wK3EQfljeUHXxYAN8uKQcvkpKBg1WM4PaVshMx7TpsBoHDtAk4c_!-1653539373"
        }
        try:
            response=requests.get(API_URL2, headers=headers, params=payload, cookies=cookies, verify=False)

            response_data=response.json()
            if response_data.get("success") == True:
                self.record_data_list=response_data.get('value',[])
                return True
            else:
                response_data.get("message", "未知錯誤")
                return False
            
        except Exception as e:
            logging.error(f"狀態碼: {response.status_code}")

    create_record=[]
    def process_procedure(self):
        wb=Workbook()
        ws=wb.active
        ws.title="紅利發放結果"
        ws.append(["帳號","活動ID","活動名稱", "紅利金額", "積分", "票卷", "票卷張數", "創建結果", "審核結果","紅利派發", "Claim_id", "紅利派發紀錄"])
        bonusAmount=10000
        bonusPointAmount=0
        #count=2
        ticketQuantity=3
        current_dir=os.path.dirname(__file__)
        yaml_path=os.path.join(current_dir,"config.yaml")
        with open(yaml_path,"r",encoding="utf-8") as f:
            config=yaml.safe_load(f)
        prmotion_id_multiple=config.get("promtion_ids",[])
        prmotion_name=config.get("promtions_name",[])
        ticket_id=config.get("ticket_id")
        testing_account=config.get("testing_account")
        confirm_result = ''
        create_result=''
        remark=""
        ticket_id_cycle=cycle(ticket_id)
        for account in testing_account:
            for promo ,name in zip(prmotion_id_multiple,prmotion_name):
                ticket=next(ticket_id_cycle)
                is_success=self.create_bonus(account,bonusAmount=bonusAmount,bonusPointAmount=bonusPointAmount,ticketId=ticket,ticketQuantity=ticketQuantity,prmotion_id=promo)
                if is_success:
                    create_result='創建紅利成功'
                    remark="成功"
                    Customerid,self.claimid = self.Search_Customer_bonus(account)
                    if self.claimid :
                        self.claimid_list.append(self.claimid)
                    if Customerid  and self.claimid :
                        is_confirm_complete=self.Confirm_Customer_bonus(Customerid)
                        if is_confirm_complete:
                            confirm_result='審核紅利成功'
                        
                        else:
                            confirm_result='審核紅利失敗'
                else:
                    create_result='創建紅利失敗'
                    remark="失敗"
                ws.append([
                        account,
                        promo,
                        name,
                        bonusAmount,
                        bonusPointAmount,
                        ticket,
                        ticketQuantity,
                        create_result,
                        confirm_result,
                        remark,
                        self.claimid,
                        "尚未比對"
                        ]
                    )
            if self.Bonus_record_page():
                Bonus_record={str(item.get("promotionClaimId") for item in self.record_data_list)}

                for row in ws.iter_rows(min_row=2,max_row=ws.max_row):
                    claimid=row[10].value
                    if claimid in Bonus_record:
                        row[11].value = "後台有紀錄"
                    else:
                        row[11].value = "後台沒紀錄"

            else:
                logging.error("無法獲取紅利記錄")
        report_path=os.path.join(current_dir,f"bonus_report_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx")
        wb.save(report_path)        
            
    
if __name__ == "__main__":

    credential = {
        "operatorName": "carrine03",
        "password": "Test@1234"
    }
    try:
        b_end=B_end(credential)
        if b_end.token:
            b_end.process_procedure()
        else:
            logging.error("登入失敗 無法取得Token")
    except Exception as e:
        logging.error(f"啟動時發生錯誤: {e}")
    
    

   